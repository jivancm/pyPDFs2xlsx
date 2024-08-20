from multiprocessing.sharedctypes import Value
from pathlib import Path, WindowsPath
from PyPDF2 import PdfReader
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook

gsDir = './EMAIL'
gsOkDir = gsDir + '/OK'
gsErrDir = gsDir + '/ERROR'
dir = Path(gsDir)
excel = "./FORMATO.xlsx"
excelNvo = "./FORMATO2.xlsx"
gsIds = []

# files = [x for x in dir if x.is_file()]

class Archivo:
    pathArchivo = None
    nombre = ""
    fecha = None
    ID = ""
    def __init__(self, pathArchivo: WindowsPath) -> None:
        if pathArchivo.suffix == ".pdf" :    
            self.pathArchivo = pathArchivo
            self.nombre = pathArchivo.name
            self.ID = self.nombre[self.nombre.find('(')+1:self.nombre.find('_')]
            fecha = self.nombre[self.nombre.find('_Date(')+6:self.nombre.find(')_', self.nombre.find('_Date(')+5)].replace('_', ':')
            fecha = fecha.replace(' (UTC)', '').replace(' (CDT)', '')
            difHoraria = fecha[len(fecha)-5:]
            fecha = fecha[:len(fecha)-6].replace(',','')
            fecha = fecha.split(' ')
            print(fecha)
            fecha[1] = ('0' if len(fecha[1])==1 else '') + fecha[1]
            fecha[4] = fecha[4].split(":")
            for i in range(3):
                fecha[4][i] = ('0' if len(fecha[4][i])==1 else '') + fecha[4][i]
            fecha = fecha[3] + "-" + fecha[2] + "-" + fecha[1] + " " + ":".join(fecha[4]) + " " + difHoraria
            try:
                self.fecha = datetime.strptime(fecha, '%Y-%b-%d %H:%M:%S %z')
            except ValueError:
                self.fecha = fecha

class PDF:
    PDF = None
    decNombre = ""
    decCurp = ""
    decRFC = ""
    decFechaPres = ""
    decFechaEntr = ""
    legible = False
    texto = ""
    decNoComprobacion = ""
    decNoTransaccion = ""
    decDependencia = ""
    decTipo = ""

    def __init__(self, ruta:WindowsPath) -> None:
        pdfInfo = None
        try:
            self.PDF = PdfReader(ruta)
            pdfInfo = self.PDF.getDocumentInfo()
        except Exception:
            print("Error")
            return None
        # print(self.PDF.documentInfo)
        if pdfInfo is None: 
            print("No se econtró información del PDF")
            return None
        if not hasattr(pdfInfo, 'creator'):
            print("No tiene la propiedad creator")
            return None
        print(pdfInfo.creator is None)
        if (pdfInfo.creator is None):
            return None
        if pdfInfo.creator[:len('Jasper')]!='Jasper':
            print("Jasper No Encontrado")
            return None
        print ("Jasper Encontrado")
        self.texto = ""
        text = ""
        for i in range(len(self.PDF.pages)) :
            try:
                text = self.PDF.pages[i].extract_text()
            except Exception:
                print("Error")
                return None
            if(len(text.replace(' ', '').rstrip()) > 0 and text.find('NO. DE COMPROBACIÓN:')>-1):
                lText = text.splitlines()
                j = lText.index('R.F.C:')
                self.decNoComprobacion = lText[j+2].replace(' ','')
                self.decNoTransaccion = lText[j+4][:lText[j+4].find('CURP:')]
                self.decRFC = lText[j+3]
                self.decCurp = lText[j+4][lText[j+4].find('CURP:')+5:]
                self.decFechaPres = self.encontrar(text, 'CIUDAD DE MÉXICO, A ')
                idx = -1
                try: 
                    idx = lText.index('PRESENTE.CIUDAD DE MÉXICO, A ' + self.decFechaPres)
                    self.decNombre = lText[idx+1][3:]
                except ValueError:
                    try:
                        idx = lText.index('CIUDAD DE MÉXICO, A ' + self.decFechaPres)
                        self.decNombre = lText[idx+1]
                    except ValueError:
                        print("No se encontró la fecha buscada, pendiente el nombre")
                if(self.decFechaPres):
                    self.decFechaPres = self.decFechaPres.replace(' DE ', '').replace('MAYO', '-05-').replace('JUNIO','-06-')
                    tmp = self.decFechaPres.split('-')
                    if(len(tmp)==3):
                        self.decFechaPres = tmp[2] + '-' + tmp[1] + '-' + tmp[0]
                self.decDependencia = self.encontrar(text, "PARA USO EXCLUSIVO EN LA ", "", True)
                self.decTipo = self.encontrar(text, "PARA USO EXCLUSIVO EN LA ", "", False, True)
                #print(text[:300])
                print("RFC: "+self.decRFC, 
                      "CURP: "+self.decCurp, 
                      self.decNoComprobacion, 
                      self.decNoTransaccion, 
                      self.decFechaPres, 
                      self.decDependencia, 
                      self.decTipo, 
                      self.decNombre)
                # print(text[:500])
        self.legible = False if (len(text.replace(' ', '').rstrip()) == 0) else True
    
    def encontrar(self, text, buscar, hasta="", siguienteLinea=False, haciaAtras=False):
        i = text.find(buscar) + (len(buscar) if not haciaAtras else 0)
        if( i == -1 ):
            return False
        if not haciaAtras:
            j = text.find("\n", i+1) if (hasta=="") else text.find(hasta, i+1)
            if (siguienteLinea):
                j = text.find("\n", j+1)
            texto = text[i:j] if (j != -1) else text[i:]
        else:
            encontrado = False
            for k in range(i-1, 0, -1):
                    if(not encontrado and text[k] == "\n") or k==0:
                        if siguienteLinea:
                            encontrado = True
                        else:
                            texto = text[k+1:i]
                            break
                    elif encontrado and text[k]=="\n":
                        texto = text[k+1:i]
                        break
        texto = texto.replace('\n', " ")
        return texto

def moverPDF(archivo):
    global gsDir
    global gsErrDir
    global gsOkDir

    anterior = Path(gsDir + '/' + archivo)
    nuevo = Path(gsOkDir + '/' + archivo)
    anterior.rename(nuevo)
    print ("Se movió de " + gsDir + '/' + archivo + " a " + gsOkDir + '/' + archivo)

i = 8
wb = load_workbook(excel)
ws = wb.active

while True:
    if ws['A' + i.__str__()].value == None:
        break
    i += 1

print("A partir de la línea " + i.__str__())

n = 1
for x in list(dir.iterdir()):
    f = Archivo(x)
    if f.ID != "":
        try:
            idx = gsIds.index(f.ID)
            #Archivo existente
            #Nuevo Archivo
        except ValueError:
            gsIds.append(f.ID)
        print(f.ID, f.fecha, type(f.fecha))
        pdf = PDF(x)
        if pdf.decCurp != "":
            ws['A' + i.__str__()].value = n
            ws['B' + i.__str__()].value = pdf.decNombre
            ws['C' + i.__str__()].value = pdf.decCurp
            if isinstance(f.fecha, str):
                ws['D' + i.__str__()].value = f.fecha
            else:
                ws['D' + i.__str__()].value = f.fecha.date()
            ws['E' + i.__str__()].value = pdf.decDependencia
            ws['F' + i.__str__()].value = pdf.decFechaPres
            ws['G' + i.__str__()].value = pdf.decRFC
            ws['H' + i.__str__()].value = f.ID
            ws['I' + i.__str__()].value = pdf.decNoComprobacion
            ws['J' + i.__str__()].value = pdf.decNoTransaccion
            ws['K' + i.__str__()].value = pdf.decTipo
            ws['L' + i.__str__()].value = f.nombre
            n += 1
            i += 1
            moverPDF(f.nombre)

wb.save(excel)
